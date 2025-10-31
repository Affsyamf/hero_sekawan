from fastapi import UploadFile
from app.utils.deps import DB 
from app.services.imports.base_import_service import BaseImportService
from io import BytesIO
import pandas as pd
import math
from openpyxl import load_workbook

from app.models import (
    Product,
    Supplier, 
    Purchasing, 
    PurchasingDetail
)

from app.utils.normalise import normalise_product_name
from app.utils.safe_parse import safe_str, safe_date, safe_number
from app.utils.cost_helper import update_avg_cost_for_products, refresh_product_avg_cost
from app.utils.event_flags import skip_cost_cache_updates
from app.utils.response import APIResponse

class LapPembelianImportService(BaseImportService):
    def __init__(self, db: DB):
        super().__init__(db)

    def _run(self, file: UploadFile):
        contents: bytes = file.file.read()
        
        wb = load_workbook(BytesIO(contents), data_only=True, keep_links=False)

        purchasings_map = {}
        count = {}
        skipped = []

        # Excel rows start at 1, Pandas index starts at 0
        HEADER_ROW = 6  # you used header=6 → means actual header row is Excel row 7
        ROW_OFFSET = HEADER_ROW + 1  # offset for correct Excel row numbers

        affected_product_ids = set()

        with skip_cost_cache_updates():

            for sheet in wb.sheetnames:
                if sheet != "AGUSTUS":
                    continue # TODO: only import AGUSTUS for now

                df = pd.read_excel(BytesIO(contents), sheet_name=sheet, header=HEADER_ROW)
                df = df.iloc[:, :-2]  # drop last 2 junk columns

                count[sheet] = 0

                for idx, row in df.iterrows():
                    excel_row_num = idx + ROW_OFFSET
                    product_name_raw = row.get("NAMA BARANG")

                    # treat None, NaN, NaT, "NAT", "NONE" as empty
                    if pd.isna(product_name_raw):
                        continue

                    product_name = safe_str(normalise_product_name(product_name_raw))
                    if not product_name or product_name in {"NAT", "NONE"}:
                        continue  # skip silently

                    # --- collect required fields ---
                    kode_supplier = safe_str(row.get("KODE SUPPLIER"))
                    tanggal = safe_date(row.get("TANGGAL"))
                    no_bukti = safe_str(row.get("NO.BUKTI"))

                    missing_cols = []
                    if not kode_supplier:
                        missing_cols.append("KODE SUPPLIER")
                    if not tanggal and not no_bukti:
                        missing_cols.append("TANGGAL/NO.BUKTI")

                    # --- log missing requireds ---
                    if missing_cols:
                        skipped.append({
                            "sheet": sheet,
                            "row": excel_row_num,
                            "reason": f"missing column(s): {', '.join(missing_cols)}",
                            "product": product_name,
                            "ppn": safe_number(row.get("PPN")),
                            "dpp": safe_number(row.get("DPP")),
                            "pph": safe_number(row.get("PPH")),
                        })
                        continue

                    # --- supplier check ---
                    supplier = self.db.query(Supplier).filter_by(code=kode_supplier).first()
                    if not supplier:
                        skipped.append({
                            "sheet": sheet,
                            "row": excel_row_num,
                            "reason": f"supplier not found: {kode_supplier}",
                            "product": product_name,
                            "ppn": safe_number(row.get("PPN")),
                            "dpp": safe_number(row.get("DPP")),
                            "pph": safe_number(row.get("PPH")),
                        })
                        continue

                    # --- purchasing key ---
                    key = (no_bukti if no_bukti else tanggal, kode_supplier)
                    if key not in purchasings_map:
                        purchasing = Purchasing(
                            date=tanggal,
                            code=no_bukti,
                            purchase_order=safe_str(row.get("NO.PO")),
                            supplier_id=supplier.id
                        )
                        self.db.add(purchasing)
                        self.db.flush()
                        purchasings_map[key] = purchasing
                    else:
                        purchasing = purchasings_map[key]

                    # --- product check ---
                    product = self.db.query(Product).filter_by(name=product_name.upper()).first()
                    if not product:
                        skipped.append({
                            "sheet": sheet,
                            "row": excel_row_num,
                            "reason": f"product not found: {product_name}",
                            "ppn": safe_number(row.get("PPN")),
                            "dpp": safe_number(row.get("DPP")),
                            "pph": safe_number(row.get("PPH")),
                        })
                        continue

                    # --- detail insert ---
                    detail = PurchasingDetail(
                        quantity=row.get("QTY") or 0,
                        price=row.get("HARGA SAT") or 0,
                        discount=row.get("POT.") or 0.0,
                        ppn=row.get("PPN") / row.get("QTY") or 0.0,
                        dpp=row.get("DPP") or 0.0,
                        pph=row.get("PPH") / row.get("QTY") or 0.0,
                        tax_no=safe_str(row.get("FAKTUR PAJAK")),
                        exchange_rate=row.get("KURS") or 0.0,
                        product_id=product.id,
                        purchasing_id=purchasing.id
                    )
                    self.db.add(detail)
                    affected_product_ids.add(product.id)
                    
                    count[sheet] += 1

            self.db.commit()

        # Bulk update the cost cache for all affected products
        if affected_product_ids:
            update_avg_cost_for_products(self.db.connection(), list(affected_product_ids))

        # refresh_product_avg_cost(self.db)

        def sanitize(obj):
            if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
                return None
            if isinstance(obj, dict):
                return {k: sanitize(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [sanitize(v) for v in obj]
            return obj

        return sanitize({
            "inserted_detail_counts": count,
            "skipped_total": len(skipped),
            "skipped_sample": skipped[:50],
        })
    
    def preview(self, file: UploadFile):
        contents: bytes = file.file.read()
        wb = load_workbook(BytesIO(contents), data_only=True, keep_links=False)

        HEADER_ROW = 6
        ROW_OFFSET = HEADER_ROW + 1

        summary = {"sheets": {}, "missing_products": set(), "missing_suppliers": set(), "skipped": []}

        for sheet in wb.sheetnames:
            if sheet != "AGUSTUS":
                continue  # TODO: only preview AGUSTUS for now

            df = pd.read_excel(BytesIO(contents), sheet_name=sheet, header=HEADER_ROW)
            df = df.iloc[:, :-2]

            rows_preview = []
            valid_rows = 0

            for idx, row in df.iterrows():
                excel_row_num = idx + ROW_OFFSET
                product_name_raw = row.get("NAMA BARANG")

                if pd.isna(product_name_raw):
                    continue

                product_name = safe_str(normalise_product_name(product_name_raw))
                if not product_name or product_name in {"NAT", "NONE"}:
                    continue

                kode_supplier = safe_str(row.get("KODE SUPPLIER"))
                tanggal = safe_date(row.get("TANGGAL"))
                no_bukti = safe_str(row.get("NO.BUKTI"))

                missing_cols = []
                if not kode_supplier:
                    missing_cols.append("KODE SUPPLIER")
                if not tanggal and not no_bukti:
                    missing_cols.append("TANGGAL/NO.BUKTI")

                if missing_cols:
                    summary["skipped"].append({
                        "sheet": sheet,
                        "row": excel_row_num,
                        "reason": f"missing column(s): {', '.join(missing_cols)}",
                        "product": product_name,
                    })
                    continue

                supplier = self.db.query(Supplier).filter_by(code=kode_supplier).first()
                if not supplier:
                    summary["missing_suppliers"].add(kode_supplier)
                    summary["skipped"].append({
                        "sheet": sheet,
                        "row": excel_row_num,
                        "reason": f"supplier not found: {kode_supplier}",
                        "product": product_name,
                    })
                    continue

                product = self.db.query(Product).filter_by(name=product_name.upper()).first()
                if not product:
                    summary["missing_products"].add(product_name)
                    summary["skipped"].append({
                        "sheet": sheet,
                        "row": excel_row_num,
                        "reason": f"product not found: {product_name}",
                    })
                    continue

                valid_rows += 1
                rows_preview.append({
                    "sheet": sheet,
                    "row": excel_row_num,
                    "supplier": kode_supplier,
                    "product": product_name,
                    "qty": safe_number(row.get("QTY")),
                    "price": safe_number(row.get("HARGA SAT")),
                    "total": round((row.get("QTY") or 0) * (row.get("HARGA SAT") or 0), 2),
                    "tanggal": tanggal.isoformat() if tanggal else None,
                    "no_bukti": no_bukti,
                })

            summary["sheets"][sheet] = {
                "valid_rows": valid_rows,
                "preview_rows": rows_preview[:30],  # limit to first 30 rows per sheet
            }

        return APIResponse.ok(
            data={
                "summary": {
                    "total_sheets": len(summary["sheets"]),
                    "missing_products": sorted(list(summary["missing_products"])),
                    "missing_suppliers": sorted(list(summary["missing_suppliers"])),
                    "skipped_total": len(summary["skipped"]),
                },
                "sheets": summary["sheets"],
                "skipped_sample": summary["skipped"][:50],
            }
        )