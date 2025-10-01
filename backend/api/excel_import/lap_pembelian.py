import pandas as pd
import json
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from io import BytesIO
import math
import re
from models.enum import LedgerLocation, LedgerRef

from utils.helpers import normalize_product_name

from db.models import Supplier, Product, Purchasing, Purchasing_Detail, Ledger


def safe_str(value):
    if pd.isna(value):
        return None
    s = str(value).strip()
    return None if s.lower() == "nan" or s == "" else s


def safe_date(value):
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.to_pydatetime()


def safe_number(value):
    """Convert NaN/inf to None, else return float/int."""
    if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
        return None
    return value


def run(contents: bytes, db: Session):
    wb = load_workbook(BytesIO(contents), data_only=True, keep_links=False)

    purchasings_map = {}
    count = {}
    skipped = []

    # Excel rows start at 1, Pandas index starts at 0
    HEADER_ROW = 6  # you used header=6 → means actual header row is Excel row 7
    ROW_OFFSET = HEADER_ROW + 1  # offset for correct Excel row numbers

    for sheet in wb.sheetnames:
        df = pd.read_excel(BytesIO(contents), sheet_name=sheet, header=HEADER_ROW)
        df = df.iloc[:, :-2]  # drop last 2 junk columns

        count[sheet] = 0

        for idx, row in df.iterrows():
            excel_row_num = idx + ROW_OFFSET
            product_name_raw = row.get("NAMA BARANG")

            # treat None, NaN, NaT, "NAT", "NONE" as empty
            if pd.isna(product_name_raw):
                continue

            product_name = safe_str(normalize_product_name(product_name_raw))
            if not product_name or product_name in {"NAT", "NONE"}:
                continue  # skip silently

            # --- collect required fields ---
            kode_supplier = safe_str(row.get("KODE SUPPLIER"))
            tanggal = safe_date(row.get("TANGGAL"))
            no_bukti = safe_str(row.get("NO.BUKTI"))

            missing_cols = []
            if not kode_supplier:
                missing_cols.append("KODE SUPPLIER")
            if not tanggal and not no_bukti:
                missing_cols.append("TANGGAL/NO.BUKTI")

            # --- log missing requireds ---
            if missing_cols:
                skipped.append({
                    "sheet": sheet,
                    "row": excel_row_num,
                    "reason": f"missing column(s): {', '.join(missing_cols)}",
                    "product": product_name,
                    "ppn": safe_number(row.get("PPN")),
                    "dpp": safe_number(row.get("DPP")),
                    "pph": safe_number(row.get("PPH")),
                })
                continue

            # --- supplier check ---
            supplier = db.query(Supplier).filter_by(code=kode_supplier).first()
            if not supplier:
                skipped.append({
                    "sheet": sheet,
                    "row": excel_row_num,
                    "reason": f"supplier not found: {kode_supplier}",
                    "product": product_name,
                    "ppn": safe_number(row.get("PPN")),
                    "dpp": safe_number(row.get("DPP")),
                    "pph": safe_number(row.get("PPH")),
                })
                continue

            # --- purchasing key ---
            key = (no_bukti if no_bukti else tanggal, kode_supplier)
            if key not in purchasings_map:
                purchasing = Purchasing(
                    date=tanggal,
                    code=no_bukti,
                    purchase_order=safe_str(row.get("NO.PO")),
                    supplier_id=supplier.id
                )
                db.add(purchasing)
                db.flush()
                purchasings_map[key] = purchasing
            else:
                purchasing = purchasings_map[key]

            # --- product check ---
            product = db.query(Product).filter_by(name=product_name.upper()).first()
            if not product:
                skipped.append({
                    "sheet": sheet,
                    "row": excel_row_num,
                    "reason": f"product not found: {product_name}",
                    "ppn": safe_number(row.get("PPN")),
                    "dpp": safe_number(row.get("DPP")),
                    "pph": safe_number(row.get("PPH")),
                })
                continue

            # --- detail insert ---
            detail = Purchasing_Detail(
                quantity=row.get("QTY") or 0,
                price=row.get("HARGA SAT") or 0,
                discount=row.get("POT.") or 0.0,
                ppn=row.get("PPN") or 0.0,
                dpp=row.get("DPP") or 0.0,
                pph=row.get("PPH") or 0.0,
                tax_no=safe_str(row.get("FAKTUR PAJAK")),
                exchange_rate=row.get("KURS") or 0.0,
                product_id=product.id,
                purchasing_id=purchasing.id
            )
            db.add(detail)
            db.flush()
            
            count[sheet] += 1
    db.commit()

    def sanitize(obj):
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        if isinstance(obj, dict):
            return {k: sanitize(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [sanitize(v) for v in obj]
        return obj

    return sanitize({
        "inserted_detail_counts": count,
        "skipped_total": len(skipped),
        "skipped_sample": skipped[:50],
    })
