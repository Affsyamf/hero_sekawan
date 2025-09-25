from __future__ import annotations
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import datetime, json
from sqlalchemy.orm import Session

HEADER_ROW = 3                       # Excel row with headers
BLOCKS = {
    "grouped": ("B", "E"),           # B3..E3
    "chemical": ("H", "K"),          # H3..K3
}
CATEGORY_KEY = "NO"                  # Column whose None/null marks a category row
GAP_TO_END = 10 

def to_iso(v):
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.isoformat()
    return v

def normalize_headers(headers: List[Any]) -> List[str]:
    """Ensure non-empty, unique string headers."""
    seen, out = set(), []
    for h in headers:
        name = (str(h).strip() if h is not None else "")
        if not name:
            name = "_"
        base, suffix = name, 2
        while name in seen:
            name = f"{base}_{suffix}"
            suffix += 1
        seen.add(name)
        out.append(name)
    return out

def _row_vals(ws, r: int, cs: int, ce: int):
    return [ws.cell(r, c).value for c in range(cs, ce + 1)]

def _looks_like_header(vals):
    return all(v is not None and str(v).strip() != "" for v in vals)

def read_block_grouped(ws, header_row: int, col_start: str, col_end: str):
    cs = column_index_from_string(col_start)
    ce = column_index_from_string(col_end)

    # headers
    headers = [ws.cell(header_row, c).value for c in range(cs, ce + 1)]
    headers = normalize_headers(headers)

    rows: List[Dict[str, Any]] = []
    seen_data = False
    r = header_row + 1

    while r <= ws.max_row:
        vals = [ws.cell(r, c).value for c in range(cs, ce + 1)]
        is_empty = all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals)

        if is_empty and seen_data:
            break  # end of this table block
        if is_empty and not seen_data:
            r += 1
            continue

        seen_data = True
        rec: Dict[str, Any] = {}
        for h, v in zip(headers, vals):
            if not h or h == "_":
                continue
            rec[h] = to_iso(v)
        rows.append(rec)
        r += 1

    # Table Batubara
    i = 0
    header_row = -1
    while i < GAP_TO_END:
        if _looks_like_header(_row_vals(ws, r + i + 1, cs, ce)):
            header_row = r + i + 1
            break
        i += 1

    coal = []
    seen_data = False
    r = header_row + 1
    while r < ws.max_row:
        vals = _row_vals(ws, r, cs, ce)
        is_empty = all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals)
        if is_empty:
            break  # end of this table block

        rec: Dict[str, Any] = {}
        for h, v in zip(headers, vals):
            if not h or h == "_":
                continue
            rec[h] = to_iso(v)
        coal.append(rec)

        r += 1

    return { 'grouped': rows, 'coal': coal }

def read_block_chem(ws, header_row: int, col_start: str, col_end: str) -> List[Dict[str, Any]]:
    cs = column_index_from_string(col_start)
    ce = column_index_from_string(col_end)

    # headers
    headers = [ws.cell(header_row, c).value for c in range(cs, ce + 1)]
    headers = normalize_headers(headers)

    rows: List[Dict[str, Any]] = []
    seen_data = False
    r = header_row + 1

    while r <= ws.max_row:
        vals = [ws.cell(r, c).value for c in range(cs, ce + 1)]
        is_empty = all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals)

        if is_empty and seen_data:
            break  # end of this table block
        if is_empty and not seen_data:
            r += 1
            continue

        seen_data = True
        rec: Dict[str, Any] = {}
        for h, v in zip(headers, vals):
            if not h or h == "_":
                continue
            rec[h] = to_iso(v)
        rows.append(rec)
        r += 1

    return rows

def group_by_category(flat_rows: List[Dict[str, Any]], category_key: str = CATEGORY_KEY) -> List[Dict[str, Any]]:
    """Turn a flat list into [{category, items:[...]}, ...] using null category_key rows as category markers."""
    grouped: List[Dict[str, Any]] = []
    current = None
    # support case-insensitive key match
    # find actual key name present in headers (e.g., "NO", "No", "no")
    actual_key = None
    if flat_rows:
        keys = flat_rows[0].keys()
        for k in keys:
            if k.lower() == category_key.lower():
                actual_key = k
                break
    if actual_key is None:
        # if not found, just wrap as a single category
        return [{"category": "All", "items": flat_rows}]

    for row in flat_rows:
        marker = row.get(actual_key, None)
        if marker in (None, "", " "):
            # category row — use the first non-empty text field as the category label
            # prefer "KELOMPOK" if present, else the first non-empty string among values
            cat = row.get("KELOMPOK") or row.get("Kelompok") or row.get("Category") or None
            if not cat:
                for v in row.values():
                    if isinstance(v, str) and v.strip():
                        cat = v.strip()
                        break
            if not cat:
                cat = "Uncategorized"
            current = {"category": cat, "items": []}
            grouped.append(current)
        else:
            if current is None:
                current = {"category": "Uncategorized", "items": []}
                grouped.append(current)
            current["items"].append(row)
    return grouped

def extract_sheet(ws) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for name, (c1, c2) in BLOCKS.items():
        if name == 'grouped':
            flat = read_block_grouped(ws, HEADER_ROW, c1, c2)
            out[name] = group_by_category(flat['grouped'], CATEGORY_KEY)
            out['coal'] = flat['coal']
        else:
            flat = read_block_chem(ws, HEADER_ROW, c1, c2)
            out[name] = flat
    return out

def run(
    contents: bytes,
    db: Session,
):
    wb = load_workbook(BytesIO(contents), data_only=True, keep_links=False)
    # optional cleanup of junk defined names
    try:
        wb.defined_names.clear()
    except Exception:
        pass

    result: Dict[str, Any] = {}
    for sheet in wb.sheetnames:
        if sheet == 'september':
            ws = wb[sheet]
            result[sheet] = extract_sheet(ws)

    # out_path = SRC.with_suffix(".json")
    # with open(out_path, "w", encoding="utf-8") as f:
    #     json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    # # tiny summary
    # for sh, blocks in result.items():
    #     print(f"{sh}: " + ", ".join(f"{b}={len(v)} cats" for b, v in blocks.items()))
    # print(f"✅ Wrote {out_path}")